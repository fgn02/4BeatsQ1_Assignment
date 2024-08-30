from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
import time
import datetime


service = Service('D:\\4BeatsQ1\\chromedriver-win64\\chromedriver.exe')
driver = webdriver.Chrome(service=service)


print("Opening Google...")
driver.get('https://www.google.com/?hl=en')
driver.implicitly_wait(20)
print("Google opened successfully.")

print("Loading Excel workbook...")
workbook_path = 'D:\\4BeatsQ1\\4BeatsQ1.xlsx'
wb = openpyxl.load_workbook(workbook_path)

current_day = datetime.datetime.now().strftime('%A')
print(f"Current day: {current_day}")

if current_day in wb.sheetnames:
    sheet = wb[current_day]
    print(f"Sheet '{current_day}' loaded successfully.")
else:
    print(f"Sheet '{current_day}' not found.")
    driver.quit()
    exit()


for row_index, row in enumerate(sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True), start=2):
    keyword = row[0]
    print(f"Row content: {keyword}")

    if keyword:
        print(f"Searching for keyword: {keyword}")

        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)

        time.sleep(2)

        WebDriverWait(driver, 10).until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//ul[@role='listbox']//li[@role='presentation']"))
        )

        lis = driver.find_elements(
            By.XPATH, "//ul[@role='listbox']//li[@role='presentation']//div[@role='option']")
        suggestions = [ele.text for ele in lis]

        if suggestions and keyword.lower() in suggestions[0].lower():
            longest_suggestion = max(suggestions, key=len)
            shortest_suggestion = min(suggestions, key=len)
            print(f"Longest suggestion: {longest_suggestion}")
            print(f"Shortest suggestion: {shortest_suggestion}")

            sheet.cell(row=row_index, column=4).value = longest_suggestion
            sheet.cell(row=row_index, column=5).value = shortest_suggestion
        else:
            print("No relevant suggestions found or keyword mismatch.")

    else:
        print("No keyword found, skipping row.")

updated_workbook_path = 'D:\\4BeatsQ1\\4BeatsQ1_updated.xlsx'
wb.save(updated_workbook_path)
print(
    f"Excel file saved with the longest and shortest suggestions at {updated_workbook_path}.")

print("Browser closed.")
driver.quit()
